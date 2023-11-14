import math
import time
from asyncio.windows_events import NULL
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from pymodbus.client import ModbusSerialClient


def createcall():
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = "measurement"
    top = ["Date", "Time", "Slave Address", "Poll Count", "Temperature", "Pressure"]
    col = 1
    for var in top:
        sheet1.cell(row=count, column=col, value=var)
        col += 1
    wb.save(name)


def filldata(vars):
    wb = openpyxl.load_workbook(name)
    sheet = wb['measurement']
    col = 1
    for var in vars:
        sheet.cell(row=count + 1, column=col, value=var)
        col += 1
    wb.save(name)


def endcall():
    end = time.time()
    totaltime = end - start
    wb = openpyxl.load_workbook(name)
    sheet2 = wb.create_sheet("sheet2")
    sheet2.title = "results"
    top = ["Completion Date", "Completion Time", "Duration", "Total Readings", "Errors count", "Error %",
           "No response count", "No response %"]
    col = 1
    for var in top:
        sheet2.cell(row=1, column=col, value=var)
        col += 1
    dati = datetime.now()
    date = dati.strftime("%d-%m-%Y")
    timee = dati.strftime("%H:%M:%S")
    m = (errorcnt / (count - 1)) * 100
    n = (norespcnt / (count - 1)) * 100
    data = [date, timee, f"{round(totaltime / 60, 2)} mins", count - 1, errorcnt, f"{m}%", norespcnt, f"{n}%"]
    col = 1
    for var in data:
        sheet2.cell(row=2, column=col, value=var)
        col += 1
    wb.save(name)
    print("Report generated!\n")


def temp():
    result = client.read_holding_registers(address=20, slave=add, count=1, unit=1)
    ans = result.registers
    return round(ans[0] * 0.1, 2)


def pressure():
    d1 = {0: 1, 1: 0.1, 2: 0.01, 3: 0.001}
    d2 = {0: "MPa", 1: "kPa", 2: "Pa", 3: "bar", 4: "mbar", 5: "kg/cm2", 6: "psi", 7: "mH20", 8: "mmH20"}
    pres = client.read_holding_registers(address=2, slave=add, count=3, unit=1)
    pf = pres.registers
    return [str(pf[2] * d1.get(pf[1])) + " " + d2.get(pf[0]), pf[2] * d1.get(pf[1])]


def main():
    p = pressure()
    te = temp()
    if p == NULL or te == NULL:
        global norespcnt
        norespcnt = norespcnt + 1
    if p[1] >= 1.5 or p[1] <= 0.5 or te <= 25 or te >= 32:
        global errorcnt
        errorcnt = errorcnt + 1
    dati = datetime.now()
    date = dati.strftime("%d-%m-%Y")
    timee = dati.strftime("%H:%M:%S")

    s = [date, timee, add, count, f"{te} C", p[0]]
    filldata(s)

    print("++++++++++++++++")
    print("poll count: ", count)
    print("date & time: ", date, timee)
    print("slave address: ", add)
    print("temperature: ", te, "C")
    print("pressure: ", p[0])
    print("++++++++++++++++\n")


# defining variables
count = 1
errorcnt = 0
norespcnt = 0

# connecting with serial port
client = ModbusSerialClient(method='rtu', port='COM2', stopbits=1, bytesize=8, baudrate=9600)
print("\nSerial port connected: ", client.connect(), "\n")

# getting slave address of the sensor
add = int(input("enter slave add: "))
frequency = int(input("Enter frequency in seconds: "))
timet = int(input("Enter test durations in mins: "))

# creating a new file
dt = datetime.now()
d = dt.strftime("%d-%m-%Y")
t = dt.strftime("%H-%M-%S")
name = f"{add}_{d}_{t}.xlsx"
createcall()

# starting the test and appending sensor readings in the newly created file
start = time.time()
print("Testing starts.... \n")
while count <= math.ceil((timet * 60) / frequency):
    main()
    time.sleep(frequency)
    count = count + 1
print("Testing completed!\n")

# generating a report
endcall()
